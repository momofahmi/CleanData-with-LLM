#!/usr/bin/env python3
"""
Checks a cleaning output against the clean truth, cell by cell.
CLEAN the original uncorrupted data
DIRTY the corrupted file given to the cleaner
CLEANED the cleaner's output
Every differing cell is classified:
INJECTED clean != dirty
REPAIRED injected and cleaned == clean
CHANGED injected, touched, but not exact
MISSED injected and cleaned == dirty
COLLATERAL clean == dirty but cleaned != clean
"""
import argparse
import sys

import numpy as np
import pandas as pd

TOL = 1e-6


def _check_cols(cols, frames):
    for name, d in frames:
        missing = [c for c in cols if c not in d.columns]
        if missing:
            sys.exit(f"column(s) {missing} not found in the {name} file")


def _align(clean, dirty, cleaned, idcols, keycols):
    frames = (("clean", clean), ("dirty", dirty), ("cleaned", cleaned))

    if idcols:
        _check_cols(idcols, frames)
        for d in (clean, dirty, cleaned):
            if d.duplicated(subset=idcols).any():
                sys.exit(f"--id {','.join(idcols)} is not unique; use --key instead, "
                         "or regenerate the clean split with make_clean_split.py")
        clean = clean.set_index(idcols)
        dirty = dirty.set_index(idcols)
        cleaned = cleaned.set_index(idcols)
        common = dirty.index.intersection(clean.index).intersection(cleaned.index)
        if not len(common):
            sys.exit("no rows in common on that key")
        return clean.loc[common], dirty.loc[common], cleaned.loc[common]

    if keycols:
        # multiset match: rows sharing a key are paired in file order
        _check_cols(keycols, frames)
        def _tagged(d):
            d = d.copy()
            d["__k"] = d[keycols].astype(str).agg("\u0001".join, axis=1)
            d["__n"] = d.groupby("__k").cumcount()
            return d.set_index(["__k", "__n"])
        clean, dirty, cleaned = _tagged(clean), _tagged(dirty), _tagged(cleaned)
        common = dirty.index.intersection(clean.index).intersection(cleaned.index)
        if not len(common):
            sys.exit("no rows in common on those key columns")
        lost = len(dirty) - len(common)
        if lost:
            print(f"note: {lost:,} dirty rows could not be matched and are ignored "
                  "(the key columns may have been corrupted).\n")
        return clean.loc[common], dirty.loc[common], cleaned.loc[common]
    if not (len(clean) == len(dirty) == len(cleaned)):
        sys.exit(f"row counts differ (clean {len(clean)}, dirty {len(dirty)}, cleaned {len(cleaned)}); "
                 "use --id <column> to align on an identifier")
    return clean.reset_index(drop=True), dirty.reset_index(drop=True), cleaned.reset_index(drop=True)


def _neq(a: pd.Series, b: pd.Series) -> pd.Series:
    """elementwise 'differs', NaN-aware, numeric-tolerant."""
    an, bn = pd.to_numeric(a, errors="coerce"), pd.to_numeric(b, errors="coerce")
    both_num = an.notna() & bn.notna()
    num_diff = both_num & ((an - bn).abs() > TOL * (1 + an.abs()))
    both_nan = a.isna() & b.isna()
    rest = ~both_num & ~both_nan
    str_diff = rest & (a.astype(str) != b.astype(str))
    one_nan = a.isna() ^ b.isna()
    return num_diff | str_diff | one_nan


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("clean"); ap.add_argument("dirty"); ap.add_argument("cleaned")
    ap.add_argument("--id", default=None,
                    help="identifier column, or comma-separated composite key (e.g. Entity,Year)")
    ap.add_argument("--key", default=None,
                    help="comma-separated columns to align on when no unique key exists "
                         "(duplicates matched in file order)")
    ap.add_argument("-o", "--xlsx", default=None, help="write a colour-coded Excel diff")
    args = ap.parse_args()

    clean = pd.read_csv(args.clean)
    dirty = pd.read_csv(args.dirty)
    cleaned = pd.read_csv(args.cleaned)
    idcols = [c.strip() for c in args.id.split(",")] if args.id else None
    keycols = [c.strip() for c in args.key.split(",")] if args.key else None
    clean, dirty, cleaned = _align(clean, dirty, cleaned, idcols, keycols)

    cols = [c for c in dirty.columns if c in clean.columns and c in cleaned.columns
            and not str(c).startswith("__")]
    print(f"aligned rows: {len(dirty):,} | compared columns: {len(cols)}\n")

    per_col, cells = {}, []
    for c in cols:
        inj = _neq(clean[c], dirty[c])
        rep = inj & ~_neq(cleaned[c], clean[c])
        mis = inj & ~_neq(cleaned[c], dirty[c])
        chg = inj & ~rep & ~mis
        col_ = ~inj & _neq(cleaned[c], clean[c])
        per_col[c] = dict(injected=int(inj.sum()), repaired=int(rep.sum()),
                          changed=int(chg.sum()), missed=int(mis.sum()),
                          collateral=int(col_.sum()))
        if args.xlsx:
            for status, mask in (("REPAIRED", rep), ("CHANGED", chg), ("MISSED", mis), ("COLLATERAL", col_)):
                for r in dirty.index[mask]:
                    cells.append(dict(row=r, column=c, status=status,
                                      clean=clean.at[r, c], dirty=dirty.at[r, c], cleaned=cleaned.at[r, c]))

    tot = {k: sum(v[k] for v in per_col.values()) for k in ("injected", "repaired", "changed", "missed", "collateral")}
    print(f"{'column':24s} {'injected':>9s} {'repaired':>9s} {'changed':>8s} {'missed':>7s} {'collateral':>10s}")
    for c, v in sorted(per_col.items(), key=lambda kv: -kv[1]["injected"]):
        if any(v.values()):
            print(f"{c[:24]:24s} {v['injected']:9,d} {v['repaired']:9,d} {v['changed']:8,d} {v['missed']:7,d} {v['collateral']:10,d}")
    print("-" * 72)
    print(f"{'TOTAL':24s} {tot['injected']:9,d} {tot['repaired']:9,d} {tot['changed']:8,d} {tot['missed']:7,d} {tot['collateral']:10,d}")
    if tot["collateral"]:
        print(f"\nWARNING: {tot['collateral']:,} healthy cells were modified (collateral damage).")

    if args.xlsx and cells:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill
        fills = {"REPAIRED": "C6EFCE", "CHANGED": "FFE4B5", "MISSED": "FFC7CE", "COLLATERAL": "E1D5F5"}
        wb = Workbook(); ws = wb.active; ws.title = "diff"
        ws.append(["row", "column", "status", "clean (truth)", "dirty (input)", "cleaned (output)"])
        cap = 50_000
        for rec in cells[:cap]:
            ws.append([str(rec[k]) for k in ("row", "column", "status", "clean", "dirty", "cleaned")])
            f = PatternFill("solid", fgColor=fills[rec["status"]])
            for cell in ws[ws.max_row]:
                cell.fill = f
        wb.save(args.xlsx)
        print(f"\nExcel diff written to {args.xlsx} ({min(len(cells), cap):,} cells"
              + (", truncated" if len(cells) > cap else "") + ")")


if __name__ == "__main__":
    main()
